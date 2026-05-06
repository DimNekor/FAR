# client/desktop/views/results_screen.py

from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.properties import ObjectProperty
from kivy.metrics import dp
from kivy.clock import Clock
from kivy.lang import Builder

from client.models.database import Database
import pandas as pd
from datetime import datetime
import os

from client.views import load_view_kv

# Загружаем KV файл
load_view_kv("results.kv")


class ResultsScreen(Screen):
    main_layout = ObjectProperty(None)
    table_container = ObjectProperty(None)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.db = Database()

    def on_enter(self):
        """Вызывается при входе на экран"""
        self.refresh_results()

    def refresh_results(self):
        """Обновление данных в таблице"""
        if not self.table_container:
            return

        self.table_container.clear_widgets()

        # Получаем данные из БД
        results = self.db.get_all_results()

        print(f"DEBUG: Found {len(results)} results in DB")
        print(f"DEBUG: DB path: {self.db.db_path}")
        
        if not results:
            no_data_label = Label(
                text="Нет данных для отображения\n\nПроведите несколько тестов, чтобы увидеть результаты",
                font_size=20,
                size_hint=(1, 1),
                halign="center",
                valign="middle",
                color=(0.5, 0.5, 0.5, 1),
            )
            self.table_container.add_widget(no_data_label)
            return

        # Создаем ScrollView для прокрутки
        scroll_view = ScrollView(
            size_hint=(1, 1),
            scroll_type=["bars", "content"],
            bar_width=dp(10),
            do_scroll_x=True,  # Горизонтальная прокрутка
            do_scroll_y=True,  # Вертикальная прокрутка
        )

        # Создаем GridLayout для таблицы
        grid = GridLayout(
            cols=11,  # Увеличили количество колонок для статуса
            size_hint=(None, None),  # Изменено для горизонтальной прокрутки
            row_force_default=True,
            row_default_height=dp(40),
            spacing=dp(2),
            padding=dp(5),
            width=dp(1200),  # Фиксированная ширина для горизонтальной прокрутки
        )

        # Заголовки таблицы с улучшенным дизайном
        headers = [
            ("ID польз.", dp(100)),
            ("Имя", dp(100)),
            ("Пол", dp(60)),
            ("Дата рег.", dp(100)),
            ("ID сессии", dp(100)),
            ("Тайминг", dp(80)),
            ("Изображение", dp(150)),
            ("Время (с)", dp(80)),
            ("Истинный", dp(80)),
            ("Предсказ.", dp(80)),
            ("Статус", dp(80)),
        ]

        for header_text, header_width in headers:
            header_label = Label(
                text=header_text,
                bold=True,
                size_hint=(None, None),
                size=(header_width, dp(35)),
                color=(1, 1, 1, 1),
                font_size=12,
                halign="center",
                valign="middle",
            )
            # Добавляем фон для заголовков
            with header_label.canvas.before:
                from kivy.graphics import Color, Rectangle

                Color(0.2, 0.2, 0.3, 1)
                header_label.bg_rect = Rectangle(
                    pos=header_label.pos, size=header_label.size
                )

            def update_header_bg(instance, value):
                instance.bg_rect.pos = instance.pos
                instance.bg_rect.size = instance.size

            header_label.bind(pos=update_header_bg, size=update_header_bg)
            grid.add_widget(header_label)

        # Данные таблицы
        for i, row in enumerate(results):
            true_class_text = "Реальное" if row["true_class"] == 1 else "Фейк"
            predicted_text = (
                "Реальное"
                if row["predicted_class"] == 1
                else "Фейк"
                if row["predicted_class"] == 0
                else "Таймаут"
            )

            is_correct = row["true_class"] == row["predicted_class"]
            status = (
                "✓ Верно"
                if is_correct
                else "✗ Ошибка"
                if row["predicted_class"] != -1
                else "⏱ Таймаут"
            )

            # Цвет строки
            if is_correct:
                bg_color = (0.9, 1, 0.9, 1)
            elif row["predicted_class"] == -1:
                bg_color = (1, 1, 0.8, 1)
            else:
                bg_color = (1, 0.9, 0.9, 1)

            values_and_widths = [
                (row["id_user"][:8] + "...", dp(100)),
                (row["participant_name"], dp(100)),
                (row["sex"], dp(60)),
                (row["registration_date"][:10], dp(100)),
                (row["id_session"][:8] + "...", dp(100)),
                (row["timing"], dp(80)),
                (row["image_name"][:20], dp(150)),
                (f"{row['reaction_time']:.2f}", dp(80)),
                (true_class_text, dp(80)),
                (predicted_text, dp(80)),
                (status, dp(80)),
            ]

            for value, width in values_and_widths:
                cell = Label(
                    text=str(value),
                    size_hint=(None, None),
                    size=(width, dp(40)),
                    color=(0, 0, 0, 1),
                    font_size=11,
                    halign="center",
                    valign="middle",
                    text_size=(width - dp(10), None),
                    shorten=True,
                )

                # Добавляем фон для ячеек
                with cell.canvas.before:
                    from kivy.graphics import Color, Rectangle

                    Color(*bg_color)
                    cell.bg_rect = Rectangle(pos=cell.pos, size=cell.size)

                def update_cell_bg(instance, value):
                    instance.bg_rect.pos = instance.pos
                    instance.bg_rect.size = instance.size

                cell.bind(pos=update_cell_bg, size=update_cell_bg)

                # Добавляем рамку
                with cell.canvas.after:
                    from kivy.graphics import Color, Line

                    Color(0.8, 0.8, 0.8, 1)
                    Line(rectangle=(cell.x, cell.y, cell.width, cell.height), width=0.5)

                grid.add_widget(cell)

        # Устанавливаем высоту грида
        grid.height = (len(results) + 1) * dp(42)

        scroll_view.add_widget(grid)
        self.table_container.add_widget(scroll_view)

    def show_export_dialog(self):
        """Показать диалог выбора места для экспорта"""
        # Проверяем, есть ли данные для экспорта
        results = self.db.get_all_results()
        if not results:
            self._show_popup("Предупреждение", "Нет данных для экспорта")
            return

        # Создаем контент для диалога
        content = BoxLayout(orientation="vertical", spacing=10, padding=10)

        # Заголовок
        content.add_widget(
            Label(
                text="Выберите папку для сохранения файла:",
                size_hint=(1, 0.1),
                bold=True,
            )
        )

        # FileChooser для выбора директории
        file_chooser = FileChooserListView(
            dirselect=True,  # Разрешаем выбор директорий
            size_hint=(1, 0.7),
            path=os.path.expanduser("~"),  # Начинаем с домашней директории
            filters=[""],  # Показываем все файлы
            multiselect=False,
        )
        content.add_widget(file_chooser)

        # Поле для имени файла
        filename_box = BoxLayout(
            orientation="horizontal", size_hint=(1, 0.1), spacing=10
        )
        filename_box.add_widget(Label(text="Имя файла:", size_hint=(0.3, 1)))

        from kivy.uix.textinput import TextInput

        default_filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        filename_input = TextInput(
            text=default_filename, size_hint=(0.7, 1), multiline=False
        )
        filename_box.add_widget(filename_input)
        content.add_widget(filename_box)

        # Кнопки
        buttons_box = BoxLayout(
            orientation="horizontal", size_hint=(1, 0.1), spacing=10
        )

        cancel_btn = Button(text="Отмена", background_color=(0.7, 0.3, 0.3, 1))

        export_btn = Button(text="Экспортировать", background_color=(0.2, 0.7, 0.3, 1))

        buttons_box.add_widget(cancel_btn)
        buttons_box.add_widget(export_btn)
        content.add_widget(buttons_box)

        popup = Popup(title="Экспорт в Excel", content=content, size_hint=(0.8, 0.8))

        def on_cancel(instance):
            popup.dismiss()

        def on_export(instance):
            selected_path = file_chooser.path
            filename = filename_input.text.strip()

            if not filename:
                self._show_popup("Ошибка", "Введите имя файла")
                return

            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            filepath = os.path.join(selected_path, filename)

            # Проверяем, не существует ли уже такой файл
            if os.path.exists(filepath):
                self._show_confirm_overwrite(popup, filepath)
            else:
                self._export_to_excel(filepath)
                popup.dismiss()

        cancel_btn.bind(on_release=on_cancel)
        export_btn.bind(on_release=on_export)

        popup.open()

    def _show_confirm_overwrite(self, parent_popup, filepath):
        """Показать подтверждение перезаписи файла"""
        content = BoxLayout(orientation="vertical", spacing=10, padding=10)
        content.add_widget(
            Label(
                text=f"Файл уже существует:\n{filepath}\n\nПерезаписать?",
                halign="center",
            )
        )

        buttons_box = BoxLayout(orientation="horizontal", spacing=10)

        no_btn = Button(text="Нет", background_color=(0.7, 0.3, 0.3, 1))
        yes_btn = Button(text="Да", background_color=(0.2, 0.7, 0.3, 1))

        buttons_box.add_widget(no_btn)
        buttons_box.add_widget(yes_btn)
        content.add_widget(buttons_box)

        confirm_popup = Popup(
            title="Подтверждение", content=content, size_hint=(0.6, 0.4)
        )

        def on_no(instance):
            confirm_popup.dismiss()

        def on_yes(instance):
            self._export_to_excel(filepath)
            confirm_popup.dismiss()
            parent_popup.dismiss()

        no_btn.bind(on_release=on_no)
        yes_btn.bind(on_release=on_yes)

        confirm_popup.open()

    def _export_to_excel(self, filepath):
        """Экспорт результатов в Excel"""
        try:
            results = self.db.get_all_results()

            # Преобразуем в DataFrame
            data = []
            for row in results:
                true_class_text = "Реальное" if row["true_class"] == 1 else "Фейк"
                predicted_text = (
                    "Реальное"
                    if row["predicted_class"] == 1
                    else "Фейк"
                    if row["predicted_class"] == 0
                    else "Таймаут"
                )
                is_correct = (
                    "Да"
                    if row["true_class"] == row["predicted_class"]
                    else "Таймаут"
                    if row["predicted_class"] == -1
                    else "Нет"
                )

                data.append(
                    {
                        "ID пользователя": row["id_user"],
                        "Имя участника": row["participant_name"],
                        "Пол": row["sex"],
                        "Дата регистрации": row["registration_date"],
                        "ID сессии": row["id_session"],
                        "Тайминг": row["timing"],
                        "Изображение": row["image_name"],
                        "Время реакции (с)": row["reaction_time"],
                        "Истинный класс": true_class_text,
                        "Предсказанный класс": predicted_text,
                        "Правильный ответ": is_correct,
                    }
                )

            df = pd.DataFrame(data)

            # Сохраняем в Excel с форматированием
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Результаты", index=False)

                # Получаем workbook и worksheet для форматирования
                workbook = writer.book
                worksheet = writer.sheets["Результаты"]

                # Стили для заголовков
                from openpyxl.styles import Font, PatternFill, Alignment

                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(
                    start_color="333366", end_color="333366", fill_type="solid"
                )
                header_alignment = Alignment(horizontal="center", vertical="center")

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Автоматически подгоняем ширину столбцов
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Выравнивание для данных
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

            self._show_popup(
                "Успех", f"Данные успешно экспортированы в файл:\n{filepath}"
            )

        except Exception as e:
            self._show_popup("Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def show_statistics(self):
        """Показать статистику в отдельном окне"""
        try:
            stats = self.db.get_statistics()

            if not stats:
                self._show_popup("Информация", "Нет данных для статистики")
                return

            # Создаем контент для popup
            content = BoxLayout(orientation="vertical", spacing=10, padding=10)

            # Заголовок
            content.add_widget(
                Label(
                    text="Статистика по сессиям",
                    bold=True,
                    size_hint=(1, 0.08),
                    font_size=16,
                )
            )

            # Создаем ScrollView для статистики
            scroll = ScrollView(size_hint=(1, 0.82))
            stats_layout = GridLayout(
                cols=3,
                size_hint_y=None,
                row_force_default=True,
                row_default_height=dp(40),
                spacing=dp(5),
            )

            # Заголовки
            headers = ["Пользователь/Сессия", "Параметр", "Значение"]
            for header in headers:
                header_label = Label(
                    text=header,
                    bold=True,
                    size_hint_y=None,
                    height=dp(35),
                    color=(1, 1, 1, 1),
                )
                with header_label.canvas.before:
                    from kivy.graphics import Color, Rectangle

                    Color(0.2, 0.2, 0.3, 1)
                    header_label.bg_rect = Rectangle(
                        pos=header_label.pos, size=header_label.size
                    )

                def update_header_bg(instance, value):
                    instance.bg_rect.pos = instance.pos
                    instance.bg_rect.size = instance.size

                header_label.bind(pos=update_header_bg, size=update_header_bg)
                stats_layout.add_widget(header_label)

            # Данные статистики
            for stat in stats:
                user_info = f"{stat['name']}\n{stat['timing']}"
                total = stat["total_images"]
                correct = stat["correct_answers"]
                accuracy = (correct / total * 100) if total > 0 else 0

                params = [
                    ("Всего изображений", str(total)),
                    ("Правильных ответов", str(correct)),
                    ("Точность", f"{accuracy:.1f}%"),
                    ("Среднее время реакции", f"{stat['avg_reaction_time']:.2f} с"),
                    ("Таймаутов", str(stat["timeouts"])),
                ]

                for i, (param, value) in enumerate(params):
                    if i == 0:
                        stats_layout.add_widget(
                            Label(
                                text=user_info,
                                size_hint_y=None,
                                height=dp(80),
                                color=(0, 0, 0.5, 1),
                                bold=True,
                            )
                        )
                    else:
                        stats_layout.add_widget(
                            Label(text="", size_hint_y=None, height=dp(40))
                        )

                    stats_layout.add_widget(
                        Label(text=param, size_hint_y=None, height=dp(40))
                    )
                    stats_layout.add_widget(
                        Label(
                            text=value,
                            size_hint_y=None,
                            height=dp(40),
                            color=(0, 0.5, 0, 1) if i == 2 else (0, 0, 0, 1),
                        )
                    )

            stats_layout.height = (len(stats) * 5 + 1) * dp(42)
            scroll.add_widget(stats_layout)
            content.add_widget(scroll)

            # Кнопка закрытия
            close_btn = Button(
                text="Закрыть", size_hint=(1, 0.1), background_color=(0.7, 0.3, 0.3, 1)
            )
            content.add_widget(close_btn)

            popup = Popup(title="Статистика", content=content, size_hint=(0.8, 0.8))

            close_btn.bind(on_release=popup.dismiss)
            popup.open()

        except Exception as e:
            self._show_popup("Ошибка", f"Ошибка при загрузке статистики:\n{str(e)}")

    def _show_popup(self, title, message):
        """Показать всплывающее окно с сообщением"""
        content = BoxLayout(orientation="vertical", spacing=10, padding=10)
        content.add_widget(Label(text=message, halign="center"))

        close_btn = Button(
            text="OK", size_hint=(1, 0.3), background_color=(0.3, 0.5, 0.7, 1)
        )
        content.add_widget(close_btn)

        popup = Popup(title=title, content=content, size_hint=(0.6, 0.4))

        close_btn.bind(on_release=popup.dismiss)
        popup.open()

    def go_to_menu(self):
        """Возврат в главное меню"""
        if self.manager:
            self.manager.current = "menu"
